import logging
import json
import requests
import os
from openpyxl import Workbook, load_workbook
from fetchemail import FetchEmail
from htmlparser import extract_htmltable
from datetime import datetime

### LOAD GLOBAL PARAMS ###
with open("config.json", "r") as f:
    config = json.load(f)
GIS_SRV = config["GLOBAL"]["GIS"]
IMAP = config["GLOBAL"]["IMAP"]
SENDER = config["GLOBAL"]["SENDER"]
SUBJECT = config["GLOBAL"]["SUBJECT"]
XL = config["GLOBAL"]["XLLOGGER"]
DOCDIR = config["GLOBAL"]["DOCDIR"]
LOGDIR = config["GLOBAL"]["LOGDIR"]

### CHECK FOLDER ###
if not os.path.isdir(DOCDIR):
    os.mkdir(DOCDIR)
if not os.path.isdir(LOGDIR):
    os.mkdir(LOGDIR)

logging.basicConfig(handlers=[logging.FileHandler(LOGDIR+"trace.log", 'a', 'utf-8')],level=logging.INFO,
                    format='%(asctime)s -- %(levelname)s -- %(module)s -- %(message)s')
logging.info("="*40)


### EXCEL LOGGER ####
try:
    wb = load_workbook(filename=LOGDIR+XL)
except OSError as e:
    logging.error("{}".format(e))
    wb = Workbook()
ws = wb.active

## Main loop ####
for account in config["ACCOUNTS"].keys():
    logging.info("Accès au compte {}".format(account))
    ID = config["ACCOUNTS"][account]["ID"]
    IMAP_PWD = config["ACCOUNTS"][account]["IMAP_PWD"]
    GIS_PWD = config["ACCOUNTS"][account]["GIS_PWD"]
    INBOX = config["ACCOUNTS"][account]["INBOX"]

    ### HTTP SESSIONS ###
    logging.info("opening http session")
    s = requests.session()
    html_payload = {"j_username": ID, "j_password": GIS_PWD}
    try:
        r = s.post(GIS_SRV, data=html_payload)
        r.raise_for_status()
    except requests.exceptions.RequestException as e:
        logging.error(e)
        continue
    # Verif contenu du cookie pour authent.
    if not r.cookies.get_dict()["decorator"] == "martinbrower":
        logging.error("Authentication has failed")
        continue

    logging.info("opening imap sessions")
    mailsrv = FetchEmail(IMAP, ID, IMAP_PWD, INBOX)
    msg_list = mailsrv.fetch_specific_messages(SENDER, SUBJECT)
    logging.info("messages en attente : {}".format(len(msg_list)))

    for msg in msg_list:
        table = extract_htmltable(msg["body"])
        uid = msg["num"].decode()
        for customer in table.keys():
            for invoice in table[customer]:
                ws.append(
                    [
                        datetime.now(),
                        account,
                        customer,
                        invoice,
                        table[customer][invoice]["montant"],
                        table[customer][invoice]["date"],
                        table[customer][invoice]["url_pdf"],
                        table[customer][invoice]["url_edi"],
                    ]
                )
                # PDF
                try:
                    r = s.get(table[customer][invoice]["url_pdf"])
                    r.raise_for_status()
                except requests.exceptions.HTTPError as e:
                    logging.error(e)
                    continue
                except requests.exceptions.RequestException as e:
                    logging.error(e)
                    continue
                size = int(len(r.content) / 1000)
                logging.info(
                    "-- {} - {}, invoice n°{} ({}kb)".format(
                        uid, customer, invoice, str(size)
                    )
                )
                # en dessous de 20KB, on considère que le download a échoué
                if size > 20:
                    open(DOCDIR + invoice + ".pdf", "wb").write(r.content)
                else:
                    logging.warning("pdf file is too small")
                # EDI
                try:
                    r = s.get(table[customer][invoice]["url_edi"])
                    r.raise_for_status()
                except requests.exceptions.HTTPError as e:
                    logging.error(e)
                    continue
                except requests.exceptions.RequestException as e:
                    logging.error(e)
                    continue
                open(DOCDIR + invoice + ".edi", "wb").write(r.content)

        # mailsrv.archive_message(msg["num"], INBOX + "/archives")

    mailsrv.close_connection()
    s.close()

wb.save(LOGDIR+XL)
wb.close()
